"""
Import Excel data into Clean Data PostgreSQL tables

Usage:
    python -m clean_data.pipelines.import_clean_data --all
    python -m clean_data.pipelines.import_clean_data --gp-file "/path/to/GP Dataset Prequin.xlsx"
"""

import os
import re
import argparse
import logging
from typing import Dict, Any, List, Optional, Generator
from datetime import datetime
import uuid

from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy import text

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# Base directory for Excel files
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
EXCEL_DIR = os.path.join(BASE_DIR, "Prequin database ")

# File paths
DEFAULT_FILES = {
    "gp": os.path.join(EXCEL_DIR, "GP Dataset Prequin.xlsx"),
    "lp": os.path.join(EXCEL_DIR, "LP Dataset Prequin.xlsx"),
    "deals": os.path.join(EXCEL_DIR, "Preqin_deals_export.xlsx"),
    "funds": os.path.join(EXCEL_DIR, "Private Market Funds.xlsx"),
}


def normalize_column_name(name: str) -> str:
    """
    Convert Excel column name to snake_case key.

    Examples:
        "FIRM NAME" -> "firm_name"
        "AUM (USD MN)" -> "aum_usd_mn"
        "YEAR EST." -> "year_est"
        "PE - Min Investment Size" -> "pe_min_investment_size"
    """
    if not name:
        return "unknown"

    # Remove leading/trailing whitespace
    name = str(name).strip()

    # Replace special characters with spaces
    name = re.sub(r'[().,\-/\\]', ' ', name)

    # Replace multiple spaces with single space
    name = re.sub(r'\s+', ' ', name)

    # Convert to lowercase and replace spaces with underscores
    name = name.lower().strip().replace(' ', '_')

    # Remove any remaining non-alphanumeric characters (except underscore)
    name = re.sub(r'[^a-z0-9_]', '', name)

    # Remove consecutive underscores
    name = re.sub(r'_+', '_', name)

    # Remove leading/trailing underscores
    name = name.strip('_')

    return name or "unknown"


def infer_data_type(values: List[Any]) -> str:
    """Infer data type from sample values."""
    non_null_values = [v for v in values if v is not None and str(v).strip()]

    if not non_null_values:
        return "string"

    # Check if all values are numeric
    numeric_count = 0
    date_count = 0

    for v in non_null_values[:100]:  # Sample first 100
        v_str = str(v).strip()

        # Check for numeric
        try:
            float(v_str.replace(",", "").replace("$", "").replace("%", ""))
            numeric_count += 1
            continue
        except (ValueError, TypeError):
            pass

        # Check for date patterns
        if re.match(r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$', v_str):
            date_count += 1
        elif re.match(r'^\d{4}[/-]\d{1,2}[/-]\d{1,2}$', v_str):
            date_count += 1

    sample_size = len(non_null_values[:100])
    if numeric_count > sample_size * 0.8:
        return "number"
    if date_count > sample_size * 0.8:
        return "date"

    return "string"


def read_excel_sheet(
    file_path: str,
    sheet_name: str,
    chunk_size: int = 5000
) -> Generator[tuple, None, None]:
    """
    Read Excel sheet in chunks using openpyxl read_only mode.

    Yields:
        (headers, chunk_rows, chunk_start_row)
    """
    logger.info(f"Opening {file_path}, sheet: {sheet_name}")

    wb = load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        logger.error(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]

    # Read headers from first row
    headers = []
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h) if h else f"column_{i}" for i, h in enumerate(header_row)]

    logger.info(f"Found {len(headers)} columns")

    # Read data rows in chunks
    chunk = []
    chunk_start_row = 2  # Data starts at row 2

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        chunk.append(row)

        if len(chunk) >= chunk_size:
            yield headers, chunk, chunk_start_row
            chunk_start_row = row_idx + 1
            chunk = []

    # Yield remaining rows
    if chunk:
        yield headers, chunk, chunk_start_row

    wb.close()


def import_sheet_to_table(
    db: Session,
    model_class,
    file_path: str,
    sheet_name: str,
    dataset_id: str,
    sheet_id: str,
    key_extractor,
    source_file: str,
    source_sheet: str,
    chunk_size: int = 5000,
    clear_existing: bool = True
) -> Dict[str, Any]:
    """
    Import a single Excel sheet into a PostgreSQL table.

    Returns:
        Dict with import statistics
    """
    from clean_data.models import ColumnMetadata

    stats = {
        "rows_imported": 0,
        "columns": 0,
        "errors": [],
        "start_time": datetime.now(),
    }

    try:
        # Clear existing data if requested
        if clear_existing:
            logger.info(f"Clearing existing data from {model_class.__tablename__}")
            db.query(model_class).delete()

            # Clear column metadata for this table
            table_name = f"{dataset_id}_{sheet_id}"
            db.query(ColumnMetadata).filter(ColumnMetadata.table_name == table_name).delete()
            db.commit()

        first_chunk = True
        all_headers = []
        normalized_headers = []
        sample_values: Dict[int, List[Any]] = {}

        for headers, rows, chunk_start in read_excel_sheet(file_path, sheet_name, chunk_size):
            if first_chunk:
                all_headers = headers
                normalized_headers = [normalize_column_name(h) for h in headers]
                stats["columns"] = len(headers)
                first_chunk = False

                # Initialize sample values collection
                sample_values = {i: [] for i in range(len(headers))}

            # Process rows in this chunk
            records = []
            for row_offset, row_values in enumerate(rows):
                row_number = chunk_start + row_offset

                # Build JSONB data dictionary
                data = {}
                for i, (header, norm_header, value) in enumerate(zip(all_headers, normalized_headers, row_values)):
                    # Store both original key and value
                    if value is not None:
                        # Convert datetime to string
                        if hasattr(value, 'isoformat'):
                            value = value.isoformat()
                        data[norm_header] = value

                        # Collect sample values for type inference
                        if len(sample_values.get(i, [])) < 100:
                            sample_values.setdefault(i, []).append(value)

                # Extract key columns
                key_values = key_extractor(data) if key_extractor else {}

                # Create record
                record = model_class(
                    id=uuid.uuid4(),
                    row_number=row_number,
                    data=data,
                    source_file=source_file,
                    source_sheet=source_sheet,
                    **{k: v for k, v in key_values.items() if hasattr(model_class, k)}
                )
                records.append(record)

            # Bulk insert chunk
            db.bulk_save_objects(records)
            db.commit()

            stats["rows_imported"] += len(records)
            logger.info(f"Imported {stats['rows_imported']} rows...")

        # Store column metadata
        table_name = f"{dataset_id}_{sheet_id}"
        logger.info(f"Storing column metadata for {table_name}")

        for i, (original, normalized) in enumerate(zip(all_headers, normalized_headers)):
            # Infer data type from samples
            samples = sample_values.get(i, [])
            data_type = infer_data_type(samples)

            # Determine default visibility (show first 12 columns by default for wide tables)
            is_visible = i < 12 if stats["columns"] > 20 else True

            col_meta = ColumnMetadata(
                id=uuid.uuid4(),
                table_name=table_name,
                column_key=normalized,
                column_name=original,
                column_index=i,
                data_type=data_type,
                is_visible_default=is_visible,
            )
            db.add(col_meta)

        db.commit()

        stats["end_time"] = datetime.now()
        stats["duration_seconds"] = (stats["end_time"] - stats["start_time"]).total_seconds()

        logger.info(f"Import complete: {stats['rows_imported']} rows, {stats['columns']} columns in {stats['duration_seconds']:.1f}s")

    except Exception as e:
        logger.error(f"Import error: {e}")
        stats["errors"].append(str(e))
        db.rollback()

    return stats


def import_all_files(
    gp_file: Optional[str] = None,
    lp_file: Optional[str] = None,
    deals_file: Optional[str] = None,
    funds_file: Optional[str] = None,
    chunk_size: int = 5000
) -> Dict[str, Any]:
    """
    Import all Excel files into PostgreSQL.
    """
    from clean_data.database import SessionLocal, init_clean_data_schema
    from clean_data.models import (
        GPFirm, GPContact, LPInvestor, LPContact, Deal, Fund, FundContact,
        KEY_COLUMN_EXTRACTORS
    )

    # Initialize schema
    init_clean_data_schema()

    db = SessionLocal()
    all_stats = {}

    try:
        # Import GP Dataset
        gp_path = gp_file or DEFAULT_FILES["gp"]
        if os.path.exists(gp_path):
            logger.info(f"\n{'='*60}\nImporting GP Dataset: {gp_path}\n{'='*60}")

            # GP Firms
            all_stats["gp_firms"] = import_sheet_to_table(
                db=db,
                model_class=GPFirm,
                file_path=gp_path,
                sheet_name="Preqin_Export",
                dataset_id="gp-dataset",
                sheet_id="firms",
                key_extractor=KEY_COLUMN_EXTRACTORS.get(GPFirm),
                source_file="GP Dataset Prequin.xlsx",
                source_sheet="Preqin_Export",
                chunk_size=chunk_size
            )

            # GP Contacts
            all_stats["gp_contacts"] = import_sheet_to_table(
                db=db,
                model_class=GPContact,
                file_path=gp_path,
                sheet_name="Contacts_Export",
                dataset_id="gp-dataset",
                sheet_id="contacts",
                key_extractor=KEY_COLUMN_EXTRACTORS.get(GPContact),
                source_file="GP Dataset Prequin.xlsx",
                source_sheet="Contacts_Export",
                chunk_size=chunk_size
            )
        else:
            logger.warning(f"GP file not found: {gp_path}")

        # Import LP Dataset
        lp_path = lp_file or DEFAULT_FILES["lp"]
        if os.path.exists(lp_path):
            logger.info(f"\n{'='*60}\nImporting LP Dataset: {lp_path}\n{'='*60}")

            # LP Investors
            all_stats["lp_investors"] = import_sheet_to_table(
                db=db,
                model_class=LPInvestor,
                file_path=lp_path,
                sheet_name="Preqin_Export",
                dataset_id="lp-dataset",
                sheet_id="investors",
                key_extractor=KEY_COLUMN_EXTRACTORS.get(LPInvestor),
                source_file="LP Dataset Prequin.xlsx",
                source_sheet="Preqin_Export",
                chunk_size=chunk_size
            )

            # LP Contacts
            all_stats["lp_contacts"] = import_sheet_to_table(
                db=db,
                model_class=LPContact,
                file_path=lp_path,
                sheet_name="Contacts_Export",
                dataset_id="lp-dataset",
                sheet_id="contacts",
                key_extractor=KEY_COLUMN_EXTRACTORS.get(LPContact),
                source_file="LP Dataset Prequin.xlsx",
                source_sheet="Contacts_Export",
                chunk_size=chunk_size
            )
        else:
            logger.warning(f"LP file not found: {lp_path}")

        # Import Deals
        deals_path = deals_file or DEFAULT_FILES["deals"]
        if os.path.exists(deals_path):
            logger.info(f"\n{'='*60}\nImporting Deals: {deals_path}\n{'='*60}")

            all_stats["deals"] = import_sheet_to_table(
                db=db,
                model_class=Deal,
                file_path=deals_path,
                sheet_name="Preqin_Export",
                dataset_id="deals-dataset",
                sheet_id="deals",
                key_extractor=KEY_COLUMN_EXTRACTORS.get(Deal),
                source_file="Preqin_deals_export.xlsx",
                source_sheet="Preqin_Export",
                chunk_size=chunk_size
            )
        else:
            logger.warning(f"Deals file not found: {deals_path}")

        # Import Funds
        funds_path = funds_file or DEFAULT_FILES["funds"]
        if os.path.exists(funds_path):
            logger.info(f"\n{'='*60}\nImporting Private Market Funds: {funds_path}\n{'='*60}")

            # Funds
            all_stats["funds"] = import_sheet_to_table(
                db=db,
                model_class=Fund,
                file_path=funds_path,
                sheet_name="Preqin_Export",
                dataset_id="funds-dataset",
                sheet_id="funds",
                key_extractor=KEY_COLUMN_EXTRACTORS.get(Fund),
                source_file="Private Market Funds.xlsx",
                source_sheet="Preqin_Export",
                chunk_size=chunk_size
            )

            # Fund Contacts
            all_stats["fund_contacts"] = import_sheet_to_table(
                db=db,
                model_class=FundContact,
                file_path=funds_path,
                sheet_name="Contacts_Export",
                dataset_id="funds-dataset",
                sheet_id="contacts",
                key_extractor=KEY_COLUMN_EXTRACTORS.get(FundContact),
                source_file="Private Market Funds.xlsx",
                source_sheet="Contacts_Export",
                chunk_size=chunk_size
            )
        else:
            logger.warning(f"Funds file not found: {funds_path}")

    finally:
        db.close()

    # Summary
    logger.info(f"\n{'='*60}\nImport Summary\n{'='*60}")
    total_rows = 0
    for table, stats in all_stats.items():
        rows = stats.get("rows_imported", 0)
        total_rows += rows
        duration = stats.get("duration_seconds", 0)
        logger.info(f"  {table}: {rows:,} rows in {duration:.1f}s")

    logger.info(f"  Total: {total_rows:,} rows")

    return all_stats


def main():
    parser = argparse.ArgumentParser(description="Import Preqin Excel data into PostgreSQL")
    parser.add_argument("--all", action="store_true", help="Import all files from default location")
    parser.add_argument("--gp-file", type=str, help="Path to GP Dataset Excel file")
    parser.add_argument("--lp-file", type=str, help="Path to LP Dataset Excel file")
    parser.add_argument("--deals-file", type=str, help="Path to Deals Excel file")
    parser.add_argument("--funds-file", type=str, help="Path to Funds Excel file")
    parser.add_argument("--chunk-size", type=int, default=5000, help="Number of rows per chunk")

    args = parser.parse_args()

    if args.all or any([args.gp_file, args.lp_file, args.deals_file, args.funds_file]):
        import_all_files(
            gp_file=args.gp_file,
            lp_file=args.lp_file,
            deals_file=args.deals_file,
            funds_file=args.funds_file,
            chunk_size=args.chunk_size
        )
    else:
        # Default: import all from default location
        import_all_files(chunk_size=args.chunk_size)


if __name__ == "__main__":
    main()
