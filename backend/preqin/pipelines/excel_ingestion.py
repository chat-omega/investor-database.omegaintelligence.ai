"""
Excel Ingestion Pipeline for Preqin Data Layer

Memory-efficient chunked reading of large Excel files (~343MB total).
Stores raw data in bronze layer with JSONB and provenance tracking.

Usage:
    python -m preqin.pipelines.excel_ingestion \
        --deals /data/Preqin_deals_export.xlsx \
        --gp "/data/GP Dataset Prequin.xlsx" \
        --lp "/data/LP Dataset Prequin.xlsx" \
        --funds "/data/Private Market Funds.xlsx"
"""

import os
import json
import uuid
import logging
from datetime import datetime
from typing import Generator, Dict, Any, Optional, List
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from preqin.database import engine, SessionLocal

logger = logging.getLogger(__name__)

# Configuration
CHUNK_SIZE = 10_000  # Rows per chunk to avoid memory issues
RUN_ID = datetime.utcnow().strftime("%Y%m%d_%H%M%S")


def ensure_checkpoint_table(session: Session) -> None:
    """
    Ensure checkpoint table exists for resumable ingestion.
    """
    create_sql = text("""
        CREATE TABLE IF NOT EXISTS preqin_bronze.ingestion_checkpoint (
            run_id TEXT NOT NULL,
            source_file TEXT NOT NULL,
            source_sheet TEXT NOT NULL DEFAULT '',
            last_row INTEGER NOT NULL,
            status TEXT DEFAULT 'in_progress',
            updated_at TIMESTAMPTZ DEFAULT NOW(),
            PRIMARY KEY (run_id, source_file, source_sheet)
        );
    """)
    session.execute(create_sql)
    session.commit()
    logger.info("Ensured checkpoint table exists")


def get_checkpoint(session: Session, run_id: str, source_file: str, source_sheet: str = "") -> int:
    """
    Get last successfully ingested row number for resume.

    Args:
        session: Database session
        run_id: Ingestion run ID
        source_file: Source file name
        source_sheet: Source sheet name

    Returns:
        Last ingested row number (0 if no checkpoint)
    """
    result = session.execute(text("""
        SELECT last_row FROM preqin_bronze.ingestion_checkpoint
        WHERE run_id = :run_id AND source_file = :source_file AND source_sheet = :source_sheet
    """), {"run_id": run_id, "source_file": source_file, "source_sheet": source_sheet})
    row = result.fetchone()
    return row[0] if row else 0


def save_checkpoint(
    session: Session,
    run_id: str,
    source_file: str,
    source_sheet: str,
    last_row: int,
    status: str = "in_progress"
) -> None:
    """
    Save ingestion checkpoint for resume capability.
    """
    session.execute(text("""
        INSERT INTO preqin_bronze.ingestion_checkpoint
            (run_id, source_file, source_sheet, last_row, status, updated_at)
        VALUES (:run_id, :source_file, :source_sheet, :last_row, :status, NOW())
        ON CONFLICT (run_id, source_file, source_sheet) DO UPDATE SET
            last_row = EXCLUDED.last_row,
            status = EXCLUDED.status,
            updated_at = NOW()
    """), {
        "run_id": run_id,
        "source_file": source_file,
        "source_sheet": source_sheet,
        "last_row": last_row,
        "status": status
    })
    session.commit()


def get_last_run_id(session: Session, source_file: str) -> Optional[str]:
    """
    Get the last incomplete run_id for a source file.

    Returns:
        run_id of last incomplete run, or None
    """
    result = session.execute(text("""
        SELECT run_id FROM preqin_bronze.ingestion_checkpoint
        WHERE source_file = :source_file AND status = 'in_progress'
        ORDER BY updated_at DESC
        LIMIT 1
    """), {"source_file": source_file})
    row = result.fetchone()
    return row[0] if row else None


def iter_excel_chunks(
    file_path: str,
    sheet_name: Optional[str] = None,
    chunk_size: int = CHUNK_SIZE,
    start_row: int = 0
) -> Generator[List[Dict[str, Any]], None, None]:
    """
    Memory-efficient iterator over Excel rows in chunks.

    Uses openpyxl read_only mode for minimal memory footprint.
    Yields lists of dictionaries (one per row).

    Args:
        file_path: Path to Excel file
        sheet_name: Optional sheet name (default: active sheet)
        chunk_size: Rows per chunk
        start_row: Row to start from (0 = beginning, for resume)
    """
    logger.info(f"Opening Excel file: {file_path}")

    # Use read_only mode for memory efficiency
    wb = load_workbook(file_path, read_only=True, data_only=True)

    # Get target sheet
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    logger.info(f"Reading sheet: {sheet_name}")

    # Get headers from first row
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    headers = [str(h).strip() if h else f"column_{i}" for i, h in enumerate(headers)]

    logger.info(f"Found {len(headers)} columns")

    if start_row > 0:
        logger.info(f"Resuming from row {start_row + 1} (skipping {start_row} rows)")

    chunk = []
    row_number = 1  # Start after header

    for row in rows:
        row_number += 1

        # Skip rows if resuming
        if start_row > 0 and row_number <= start_row:
            continue

        # Convert row to dictionary
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                # Handle various data types
                if value is not None:
                    if hasattr(value, 'isoformat'):  # datetime
                        row_dict[headers[i]] = value.isoformat()
                    elif isinstance(value, (int, float)):
                        row_dict[headers[i]] = value
                    else:
                        stripped = str(value).strip()
                        # Convert empty/whitespace-only strings to None
                        row_dict[headers[i]] = stripped if stripped else None
                else:
                    row_dict[headers[i]] = None
        
        # Add provenance
        row_dict["_source_row_number"] = row_number
        
        chunk.append(row_dict)
        
        if len(chunk) >= chunk_size:
            logger.info(f"Yielding chunk of {len(chunk)} rows (rows {row_number - len(chunk) + 1}-{row_number})")
            yield chunk
            chunk = []
    
    # Yield remaining rows
    if chunk:
        logger.info(f"Yielding final chunk of {len(chunk)} rows")
        yield chunk
    
    wb.close()
    logger.info(f"Finished reading {row_number - 1} rows from {file_path}")


def insert_bronze_batch(
    session: Session,
    table_name: str,
    records: List[Dict[str, Any]],
    source_file: str,
    source_sheet: str,
    run_id: str
) -> int:
    """
    Batch insert records into bronze table.
    
    Uses PostgreSQL JSONB for flexible schema.
    Returns number of records inserted.
    """
    if not records:
        return 0
    
    # Prepare batch insert
    values = []
    for record in records:
        row_number = record.pop("_source_row_number", None)
        values.append({
            "id": str(uuid.uuid4()),
            "raw_data": json.dumps(record),
            "source_file": source_file,
            "source_sheet": source_sheet,
            "source_row_number": row_number,
            "run_id": run_id,
            "created_at": datetime.utcnow().isoformat()
        })
    
    # Build insert statement - use CAST instead of :: for batch compatibility
    insert_sql = text(f"""
        INSERT INTO preqin_bronze.{table_name}
        (id, raw_data, source_file, source_sheet, source_row_number, run_id, created_at)
        VALUES (
            CAST(:id AS UUID),
            CAST(:raw_data AS JSONB),
            :source_file,
            :source_sheet,
            :source_row_number,
            :run_id,
            CAST(:created_at AS TIMESTAMPTZ)
        )
    """)

    # Execute as individual inserts to avoid batch parameter issues
    for record in values:
        session.execute(insert_sql, record)
    session.commit()

    return len(values)


def ensure_bronze_table(session: Session, table_name: str) -> None:
    """
    Ensure bronze table exists with proper schema.
    """
    create_sql = text(f"""
        CREATE TABLE IF NOT EXISTS preqin_bronze.{table_name} (
            id UUID PRIMARY KEY,
            raw_data JSONB NOT NULL,
            source_file TEXT,
            source_sheet TEXT,
            source_row_number INTEGER,
            run_id TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            processed_at TIMESTAMPTZ,
            processing_errors JSONB
        );
        
        CREATE INDEX IF NOT EXISTS idx_{table_name}_run_id 
        ON preqin_bronze.{table_name} (run_id);
        
        CREATE INDEX IF NOT EXISTS idx_{table_name}_raw_data 
        ON preqin_bronze.{table_name} USING gin (raw_data);
    """)
    
    session.execute(create_sql)
    session.commit()
    logger.info(f"Ensured bronze table exists: preqin_bronze.{table_name}")


def ingest_deals(
    file_path: str,
    sheet_name: Optional[str] = None,
    run_id: str = RUN_ID,
    resume: bool = False
) -> Dict[str, Any]:
    """
    Ingest Preqin deals export.

    Expected file: Preqin_deals_export.xlsx (~139MB, 834K rows)

    Args:
        file_path: Path to Excel file
        sheet_name: Optional sheet name
        run_id: Ingestion run ID
        resume: If True, resume from last checkpoint
    """
    table_name = "deals_raw"
    source_file = Path(file_path).name
    sheet = sheet_name or "Sheet1"

    logger.info(f"Starting deals ingestion from {file_path}")

    with SessionLocal() as session:
        ensure_bronze_table(session, table_name)
        ensure_checkpoint_table(session)

        # Check for resume
        start_row = 0
        if resume:
            # Find last incomplete run for this file
            last_run = get_last_run_id(session, source_file)
            if last_run:
                run_id = last_run
                start_row = get_checkpoint(session, run_id, source_file, sheet)
                logger.info(f"Resuming run {run_id} from row {start_row}")
            else:
                logger.info("No incomplete run found, starting fresh")

        total_rows = start_row  # Count from checkpoint
        chunk_count = 0
        last_row = start_row

        for chunk in iter_excel_chunks(file_path, sheet_name, start_row=start_row):
            inserted = insert_bronze_batch(
                session, table_name, chunk,
                source_file, sheet, run_id
            )
            # Track last row from the batch
            if chunk:
                last_row = chunk[-1].get("_source_row_number", last_row + len(chunk))

            total_rows += inserted
            chunk_count += 1
            logger.info(f"Deals: Inserted chunk {chunk_count}, total rows: {total_rows}")

            # Save checkpoint after each chunk
            save_checkpoint(session, run_id, source_file, sheet, last_row)

        # Mark complete
        save_checkpoint(session, run_id, source_file, sheet, last_row, status="completed")
    
    result = {
        "table": f"preqin_bronze.{table_name}",
        "file": source_file,
        "total_rows": total_rows,
        "chunks": chunk_count,
        "run_id": run_id
    }
    logger.info(f"Deals ingestion complete: {result}")
    return result


def ingest_gp_firms(
    file_path: str,
    run_id: str = RUN_ID,
    resume: bool = False
) -> Dict[str, Any]:
    """
    Ingest GP Dataset from Preqin.

    Expected file: GP Dataset Prequin.xlsx (~75MB)
    Contains multiple sheets: Firm Profile, Funds, Contacts, etc.
    """
    logger.info(f"Starting GP firms ingestion from {file_path}")
    source_file = Path(file_path).name

    # Load workbook to get sheet names
    wb = load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    logger.info(f"Found sheets in GP file: {sheet_names}")

    results = {}

    with SessionLocal() as session:
        ensure_checkpoint_table(session)

        # Check for resume
        if resume:
            last_run = get_last_run_id(session, source_file)
            if last_run:
                run_id = last_run
                logger.info(f"Resuming run {run_id}")

        for sheet_name in sheet_names:
            # Determine table name based on sheet
            table_suffix = sheet_name.lower().replace(" ", "_").replace("-", "_")
            table_name = f"gp_{table_suffix}_raw"

            ensure_bronze_table(session, table_name)

            # Check checkpoint for this sheet
            start_row = 0
            if resume:
                start_row = get_checkpoint(session, run_id, source_file, sheet_name)
                if start_row > 0:
                    logger.info(f"Resuming {sheet_name} from row {start_row}")

            total_rows = start_row
            chunk_count = 0
            last_row = start_row

            try:
                for chunk in iter_excel_chunks(file_path, sheet_name, start_row=start_row):
                    inserted = insert_bronze_batch(
                        session, table_name, chunk,
                        source_file, sheet_name, run_id
                    )
                    if chunk:
                        last_row = chunk[-1].get("_source_row_number", last_row + len(chunk))
                    total_rows += inserted
                    chunk_count += 1
                    logger.info(f"GP {sheet_name}: Inserted chunk {chunk_count}, total rows: {total_rows}")
                    save_checkpoint(session, run_id, source_file, sheet_name, last_row)

                save_checkpoint(session, run_id, source_file, sheet_name, last_row, status="completed")
            except Exception as e:
                logger.error(f"Error processing GP sheet {sheet_name}: {e}")
                results[sheet_name] = {"error": str(e)}
                continue

            results[sheet_name] = {
                "table": f"preqin_bronze.{table_name}",
                "total_rows": total_rows,
                "chunks": chunk_count
            }

    logger.info(f"GP ingestion complete: {len(results)} sheets processed")
    return {"file": source_file, "sheets": results, "run_id": run_id}


def ingest_lp_firms(
    file_path: str,
    run_id: str = RUN_ID,
    resume: bool = False
) -> Dict[str, Any]:
    """
    Ingest LP Dataset from Preqin.

    Expected file: LP Dataset Prequin.xlsx (~67MB)
    Contains multiple sheets: Firm Profile, Fund Investments, Contacts, etc.
    """
    logger.info(f"Starting LP firms ingestion from {file_path}")
    source_file = Path(file_path).name

    # Load workbook to get sheet names
    wb = load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    logger.info(f"Found sheets in LP file: {sheet_names}")

    results = {}

    with SessionLocal() as session:
        ensure_checkpoint_table(session)

        # Check for resume
        if resume:
            last_run = get_last_run_id(session, source_file)
            if last_run:
                run_id = last_run
                logger.info(f"Resuming run {run_id}")

        for sheet_name in sheet_names:
            # Determine table name based on sheet
            table_suffix = sheet_name.lower().replace(" ", "_").replace("-", "_")
            table_name = f"lp_{table_suffix}_raw"

            ensure_bronze_table(session, table_name)

            # Check checkpoint for this sheet
            start_row = 0
            if resume:
                start_row = get_checkpoint(session, run_id, source_file, sheet_name)
                if start_row > 0:
                    logger.info(f"Resuming {sheet_name} from row {start_row}")

            total_rows = start_row
            chunk_count = 0
            last_row = start_row

            try:
                for chunk in iter_excel_chunks(file_path, sheet_name, start_row=start_row):
                    inserted = insert_bronze_batch(
                        session, table_name, chunk,
                        source_file, sheet_name, run_id
                    )
                    if chunk:
                        last_row = chunk[-1].get("_source_row_number", last_row + len(chunk))
                    total_rows += inserted
                    chunk_count += 1
                    logger.info(f"LP {sheet_name}: Inserted chunk {chunk_count}, total rows: {total_rows}")
                    save_checkpoint(session, run_id, source_file, sheet_name, last_row)

                save_checkpoint(session, run_id, source_file, sheet_name, last_row, status="completed")
            except Exception as e:
                logger.error(f"Error processing LP sheet {sheet_name}: {e}")
                results[sheet_name] = {"error": str(e)}
                continue

            results[sheet_name] = {
                "table": f"preqin_bronze.{table_name}",
                "total_rows": total_rows,
                "chunks": chunk_count
            }

    logger.info(f"LP ingestion complete: {len(results)} sheets processed")
    return {"file": source_file, "sheets": results, "run_id": run_id}


def ingest_funds(
    file_path: str,
    run_id: str = RUN_ID,
    resume: bool = False
) -> Dict[str, Any]:
    """
    Ingest Private Market Funds from Preqin.

    Expected file: Private Market Funds.xlsx (~62MB)
    """
    logger.info(f"Starting funds ingestion from {file_path}")
    source_file = Path(file_path).name

    # Load workbook to get sheet names
    wb = load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    logger.info(f"Found sheets in Funds file: {sheet_names}")

    results = {}

    with SessionLocal() as session:
        ensure_checkpoint_table(session)

        # Check for resume
        if resume:
            last_run = get_last_run_id(session, source_file)
            if last_run:
                run_id = last_run
                logger.info(f"Resuming run {run_id}")

        for sheet_name in sheet_names:
            # Determine table name based on sheet
            table_suffix = sheet_name.lower().replace(" ", "_").replace("-", "_")
            table_name = f"funds_{table_suffix}_raw"

            ensure_bronze_table(session, table_name)

            # Check checkpoint for this sheet
            start_row = 0
            if resume:
                start_row = get_checkpoint(session, run_id, source_file, sheet_name)
                if start_row > 0:
                    logger.info(f"Resuming {sheet_name} from row {start_row}")

            total_rows = start_row
            chunk_count = 0
            last_row = start_row

            try:
                for chunk in iter_excel_chunks(file_path, sheet_name, start_row=start_row):
                    inserted = insert_bronze_batch(
                        session, table_name, chunk,
                        source_file, sheet_name, run_id
                    )
                    if chunk:
                        last_row = chunk[-1].get("_source_row_number", last_row + len(chunk))
                    total_rows += inserted
                    chunk_count += 1
                    logger.info(f"Funds {sheet_name}: Inserted chunk {chunk_count}, total rows: {total_rows}")
                    save_checkpoint(session, run_id, source_file, sheet_name, last_row)

                save_checkpoint(session, run_id, source_file, sheet_name, last_row, status="completed")
            except Exception as e:
                logger.error(f"Error processing Funds sheet {sheet_name}: {e}")
                results[sheet_name] = {"error": str(e)}
                continue

            results[sheet_name] = {
                "table": f"preqin_bronze.{table_name}",
                "total_rows": total_rows,
                "chunks": chunk_count
            }

    logger.info(f"Funds ingestion complete: {len(results)} sheets processed")
    return {"file": source_file, "sheets": results, "run_id": run_id}


def ingest_all_files(
    deals_path: Optional[str] = None,
    gp_path: Optional[str] = None,
    lp_path: Optional[str] = None,
    funds_path: Optional[str] = None,
    data_dir: Optional[str] = None,
    resume: bool = False
) -> Dict[str, Any]:
    """
    Ingest all Preqin Excel files.

    Either provide individual paths or a data_dir containing:
    - Preqin_deals_export.xlsx
    - GP Dataset Prequin.xlsx
    - LP Dataset Prequin.xlsx
    - Private Market Funds.xlsx

    Args:
        deals_path: Path to deals file
        gp_path: Path to GP file
        lp_path: Path to LP file
        funds_path: Path to funds file
        data_dir: Directory containing all files
        resume: If True, resume from last checkpoint
    """
    run_id = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    results = {"run_id": run_id, "files": {}, "resumed": resume}

    # Resolve paths from data_dir if not provided individually
    if data_dir:
        data_path = Path(data_dir)
        deals_path = deals_path or str(data_path / "Preqin_deals_export.xlsx")
        gp_path = gp_path or str(data_path / "GP Dataset Prequin.xlsx")
        lp_path = lp_path or str(data_path / "LP Dataset Prequin.xlsx")
        funds_path = funds_path or str(data_path / "Private Market Funds.xlsx")

    # Ingest each file if path exists
    if deals_path and Path(deals_path).exists():
        logger.info("=" * 60)
        logger.info("INGESTING DEALS")
        logger.info("=" * 60)
        results["files"]["deals"] = ingest_deals(deals_path, run_id=run_id, resume=resume)
    else:
        logger.warning(f"Deals file not found: {deals_path}")

    if gp_path and Path(gp_path).exists():
        logger.info("=" * 60)
        logger.info("INGESTING GP FIRMS")
        logger.info("=" * 60)
        results["files"]["gp"] = ingest_gp_firms(gp_path, run_id=run_id, resume=resume)
    else:
        logger.warning(f"GP file not found: {gp_path}")

    if lp_path and Path(lp_path).exists():
        logger.info("=" * 60)
        logger.info("INGESTING LP FIRMS")
        logger.info("=" * 60)
        results["files"]["lp"] = ingest_lp_firms(lp_path, run_id=run_id, resume=resume)
    else:
        logger.warning(f"LP file not found: {lp_path}")

    if funds_path and Path(funds_path).exists():
        logger.info("=" * 60)
        logger.info("INGESTING FUNDS")
        logger.info("=" * 60)
        results["files"]["funds"] = ingest_funds(funds_path, run_id=run_id, resume=resume)
    else:
        logger.warning(f"Funds file not found: {funds_path}")
    
    # Summary
    total_rows = 0
    for file_result in results["files"].values():
        if "total_rows" in file_result:
            total_rows += file_result["total_rows"]
        elif "sheets" in file_result:
            for sheet_result in file_result["sheets"].values():
                if isinstance(sheet_result, dict) and "total_rows" in sheet_result:
                    total_rows += sheet_result["total_rows"]
    
    results["summary"] = {
        "total_rows_ingested": total_rows,
        "files_processed": len(results["files"])
    }
    
    logger.info("=" * 60)
    logger.info(f"INGESTION COMPLETE: {total_rows:,} rows from {len(results['files'])} files")
    logger.info("=" * 60)
    
    return results


# =============================================================================
# CLI Entry Point
# =============================================================================

if __name__ == "__main__":
    import argparse
    
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
    )
    
    parser = argparse.ArgumentParser(description="Ingest Preqin Excel files to PostgreSQL")
    parser.add_argument("--deals", help="Path to Preqin_deals_export.xlsx")
    parser.add_argument("--gp", help="Path to GP Dataset Prequin.xlsx")
    parser.add_argument("--lp", help="Path to LP Dataset Prequin.xlsx")
    parser.add_argument("--funds", help="Path to Private Market Funds.xlsx")
    parser.add_argument("--data-dir", help="Directory containing all Preqin Excel files")
    parser.add_argument("--chunk-size", type=int, default=CHUNK_SIZE,
                        help=f"Rows per chunk (default: {CHUNK_SIZE})")
    parser.add_argument("--resume", action="store_true",
                        help="Resume from last checkpoint (if interrupted)")

    args = parser.parse_args()

    # Update global chunk size if specified
    if args.chunk_size != CHUNK_SIZE:
        CHUNK_SIZE = args.chunk_size
        logger.info(f"Using chunk size: {CHUNK_SIZE}")

    if args.resume:
        logger.info("Resume mode enabled - will continue from last checkpoint")

    # Run ingestion
    result = ingest_all_files(
        deals_path=args.deals,
        gp_path=args.gp,
        lp_path=args.lp,
        funds_path=args.funds,
        data_dir=args.data_dir,
        resume=args.resume
    )
    
    # Print summary as JSON
    print(json.dumps(result, indent=2, default=str))
